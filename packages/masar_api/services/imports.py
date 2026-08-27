"""رفع الجدول الأسبوعي والتحقق منه (§9).

القاعدة الحاكمة: **رفع الملف بنجاح لا يعني أن الخطة قابلة للتنفيذ.** لذلك
التحقق هنا على مستويين:

1. **تحقق بنيوي** — الحقول والصيغ والمراجع والتكرار.
2. **فحص جدوى مبدئي** — هل يمكن رياضيًا تنفيذ هذه الشحنة أصلًا؟ (نافذة
   مقلوبة، SLA قبل أبكر تسليم ممكن، مسافة تتجاوز الوردية…). يستخدم نفس
   دالة ``shipment_prescreen`` التي يستخدمها المحرك، فلا يختلف الحكم بين
   شاشة الرفع وشاشة التخطيط.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import re
import unicodedata
import uuid
from dataclasses import dataclass, field
from typing import Any, Iterable

import pgwire
from masar_core.constants import AuditAction, ImportStatus, ServiceType, ShipmentStatus
from masar_core.errors import Conflict, NotFound, ValidationError
from masar_core.timeutil import (
    DEFAULT_TZ,
    combine_local,
    normalize_digits,
    parse_date,
    parse_time,
    to_utc,
)
from masar_db.driver import SecurityContext, session, transaction

from . import audit, settings as settings_service

# ------------------------------------------------------ أعمدة القالب -------


@dataclass(frozen=True, slots=True)
class TemplateColumn:
    key: str
    label_ar: str
    required: bool
    example: str
    aliases: tuple[str, ...] = ()
    note_ar: str = ""


TEMPLATE_COLUMNS: tuple[TemplateColumn, ...] = (
    TemplateColumn("external_reference", "رقم الشحنة الخارجي", False, "REQ-100234",
                   ("external_reference", "رقم الطلب", "المرجع", "reference")),
    TemplateColumn("service_date", "تاريخ الخدمة", True, "2026-09-06",
                   ("service_date", "التاريخ", "تاريخ التنفيذ", "date"),
                   "الصيغ المقبولة: YYYY-MM-DD أو DD/MM/YYYY"),
    TemplateColumn("service_type", "نوع الخدمة", False, "ROUTINE",
                   ("service_type", "الخدمة", "نوع الطلب"),
                   "ROUTINE أو URGENT أو STAT أو RETURN"),
    TemplateColumn("hub_code", "رمز مركز الانطلاق", True, "H-RYD-1",
                   ("hub_code", "مركز الانطلاق", "المركز", "hub")),
    TemplateColumn("pickup_facility_code", "رمز جهة الالتقاط", True, "PHC-RYD-014",
                   ("pickup_facility_code", "جهة الالتقاط", "الجهة", "pickup")),
    TemplateColumn("pickup_time", "موعد الالتقاط", True, "07:30",
                   ("pickup_time", "وقت الالتقاط", "موعد الالتقاط", "pickup_at"),
                   "وقت محلي HH:MM — تُبنى النافذة حوله بالسماحية المعتمدة"),
    TemplateColumn("pickup_window_from", "بداية نافذة الالتقاط", False, "07:15",
                   ("pickup_window_from", "بداية النافذة"),
                   "اتركه فارغًا لاستخدام السماحية الافتراضية"),
    TemplateColumn("pickup_window_to", "نهاية نافذة الالتقاط", False, "07:45",
                   ("pickup_window_to", "نهاية النافذة")),
    TemplateColumn("dropoff_facility_code", "رمز جهة التسليم", True, "LAB-RYD-01",
                   ("dropoff_facility_code", "جهة التسليم", "المختبر", "dropoff")),
    TemplateColumn("sla_time", "الموعد النهائي للتسليم", True, "11:00",
                   ("sla_time", "SLA", "موعد التسليم", "sla_deadline"),
                   "وقت محلي في نفس اليوم، أو تاريخ ووقت كاملان"),
    TemplateColumn("piece_count", "عدد القطع", False, "3",
                   ("piece_count", "عدد العينات", "الكمية", "pieces")),
    TemplateColumn("sample_types", "نوع العينات", False, "دم، بول",
                   ("sample_types", "أنواع العينات", "العينات")),
    TemplateColumn("temperature_mode", "نطاق الحرارة", False, "CHILLED",
                   ("temperature_mode", "الحرارة", "نطاق الحفظ"),
                   "AMBIENT أو CHILLED أو FROZEN أو DEEP_FROZEN أو CONTROLLED"),
    TemplateColumn("pickup_contact_name", "مسؤول الالتقاط", False, "أحمد العتيبي",
                   ("pickup_contact_name", "مسؤول التسليم من الجهة")),
    TemplateColumn("pickup_contact_phone", "جوال مسؤول الالتقاط", False, "0500000000",
                   ("pickup_contact_phone", "جوال الالتقاط")),
    TemplateColumn("dropoff_contact_name", "مسؤول الاستلام", False, "سارة القحطاني",
                   ("dropoff_contact_name", "مستلم العينات")),
    TemplateColumn("dropoff_contact_phone", "جوال مسؤول الاستلام", False, "0511111111",
                   ("dropoff_contact_phone", "جوال التسليم")),
    TemplateColumn("notes", "ملاحظات", False, "", ("notes", "ملاحظة", "بيان")),
)

COLUMN_INDEX = {column.key: column for column in TEMPLATE_COLUMNS}


# --------------------------------------------------------- أخطاء الصف ------

@dataclass(slots=True)
class RowIssue:
    column: str | None
    code: str
    message_ar: str
    severity: str = "ERROR"   # ERROR | WARNING

    def to_dict(self) -> dict[str, Any]:
        return {
            "column": self.column,
            "column_label_ar": COLUMN_INDEX[self.column].label_ar if self.column in COLUMN_INDEX else self.column,
            "code": self.code,
            "message_ar": self.message_ar,
            "severity": self.severity,
        }


@dataclass(slots=True)
class ParsedRow:
    row_number: int
    raw: dict[str, Any]
    normalized: dict[str, Any] = field(default_factory=dict)
    issues: list[RowIssue] = field(default_factory=list)
    dedupe_key: str | None = None

    @property
    def is_valid(self) -> bool:
        return not any(issue.severity == "ERROR" for issue in self.issues)

    @property
    def warnings(self) -> list[RowIssue]:
        return [issue for issue in self.issues if issue.severity == "WARNING"]

    @property
    def errors(self) -> list[RowIssue]:
        return [issue for issue in self.issues if issue.severity == "ERROR"]


# ------------------------------------------------------- قراءة الملفات -----

def _normalize_header(text: str) -> str:
    text = unicodedata.normalize("NFKC", str(text or "")).strip().lower()
    text = re.sub(r"[ً-ْـ]", "", text)   # تشكيل وتطويل
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ة", "ه")
    return re.sub(r"[\s_\-/]+", "", text)


def suggest_mapping(headers: list[str]) -> dict[str, str | None]:
    """يقترح مطابقة الأعمدة عند اختلاف أسمائها (§9 خطوة ٤)."""
    normalized = {_normalize_header(header): header for header in headers}
    mapping: dict[str, str | None] = {}
    for column in TEMPLATE_COLUMNS:
        candidates = (column.label_ar, column.key, *column.aliases)
        found: str | None = None
        for candidate in candidates:
            key = _normalize_header(candidate)
            if key in normalized:
                found = normalized[key]
                break
        if found is None:
            for norm_header, original in normalized.items():
                if any(_normalize_header(c) in norm_header for c in candidates if len(c) > 3):
                    found = original
                    break
        mapping[column.key] = found
    return mapping


def read_table(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """يقرأ CSV أو XLSX ويعيد (الترويسات، الصفوف كقواميس)."""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(content)
    return _read_csv(content)


def _read_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "windows-1256"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValidationError("تعذر فك ترميز الملف — احفظه بترميز UTF-8")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [h.strip() for h in (reader.fieldnames or [])]
    if not headers:
        raise ValidationError("الملف بلا صف ترويسة")
    rows = [{(k or "").strip(): v for k, v in row.items()} for row in reader]
    return headers, rows


def _read_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise ValidationError(
            "قراءة ملفات Excel غير متاحة في هذا الخادم — ارفع الملف بصيغة CSV"
        ) from exc

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        raise ValidationError("الملف فارغ") from None
    headers = [str(cell).strip() if cell is not None else f"عمود{i+1}"
               for i, cell in enumerate(header_row)]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        if values is None or all(value in (None, "") for value in values):
            continue
        rows.append({
            headers[i]: values[i] if i < len(values) else None
            for i in range(len(headers))
        })
    workbook.close()
    return headers, rows


def build_template_csv() -> bytes:
    """قالب CSV جاهز للتنزيل (§9 خطوة ١)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([column.label_ar for column in TEMPLATE_COLUMNS])
    writer.writerow([column.example for column in TEMPLATE_COLUMNS])
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def build_template_xlsx() -> bytes | None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:  # pragma: no cover
        return None

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "جدول النقل الأسبوعي"
    sheet.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="1F5C4A")
    header_font = Font(color="FFFFFF", bold=True)
    for index, column in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=column.label_ar)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = max(16, len(column.label_ar) + 6)
        sheet.cell(row=2, column=index, value=column.example)
        note = column.note_ar or ("حقل إلزامي" if column.required else "حقل اختياري")
        sheet.cell(row=3, column=index, value=note).font = Font(size=9, italic=True)

    sheet.freeze_panes = "A2"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------ التحقق -------

def _lookup_reference_data(conn: pgwire.Connection) -> dict[str, Any]:
    hubs = conn.fetch_all(
        "SELECT id::text AS id, code, name_ar, region_id::text AS region_id, "
        "city_id::text AS city_id, lat, lon FROM hubs WHERE is_active"
    )
    facilities = conn.fetch_all(
        "SELECT id::text AS id, code, name_ar, facility_type, lat, lon, "
        "service_minutes, region_id::text AS region_id, city_id::text AS city_id, "
        "default_hub_id::text AS default_hub_id, contact_name, contact_phone, address "
        "FROM facilities WHERE is_active AND voided_at IS NULL"
    )
    return {
        "hubs_by_code": {row["code"].strip().upper(): row for row in hubs},
        "facilities_by_code": {row["code"].strip().upper(): row for row in facilities},
    }


def validate_rows(
    conn: pgwire.Connection,
    rows: list[dict[str, Any]],
    mapping: dict[str, str | None],
    *,
    timezone_name: str = DEFAULT_TZ,
) -> list[ParsedRow]:
    """يحول الصفوف الخام إلى صفوف مُتحقق منها مع أخطاء محددة بالصف والعمود."""
    reference = _lookup_reference_data(conn)
    settings_cache: dict[str, dict[str, Any]] = {}
    parsed: list[ParsedRow] = []
    seen_keys: dict[str, int] = {}

    for offset, raw in enumerate(rows):
        row_number = offset + 2  # الصف ١ ترويسة
        row = ParsedRow(row_number=row_number, raw={k: _plain(v) for k, v in raw.items()})

        def value_of(key: str) -> Any:
            source = mapping.get(key)
            if not source:
                return None
            item = raw.get(source)
            if isinstance(item, str):
                item = item.strip()
            return item if item not in ("", None) else None

        def add_error(column: str | None, code: str, message: str) -> None:
            row.issues.append(RowIssue(column, code, message))

        def add_warning(column: str | None, code: str, message: str) -> None:
            row.issues.append(RowIssue(column, code, message, "WARNING"))

        # --------------------------------------------- الحقول المطلوبة ---
        for column in TEMPLATE_COLUMNS:
            if column.required and value_of(column.key) is None:
                add_error(column.key, "MISSING_REQUIRED",
                          f"الحقل «{column.label_ar}» مطلوب وفارغ")

        if not row.is_valid:
            parsed.append(row)
            continue

        # ------------------------------------------------------ التاريخ --
        try:
            service_date = parse_date(value_of("service_date"), field="تاريخ الخدمة")
        except ValidationError as exc:
            add_error("service_date", "BAD_DATE", exc.message)
            parsed.append(row)
            continue

        # ------------------------------------------------ مركز الانطلاق --
        hub_code = str(value_of("hub_code") or "").strip().upper()
        hub = reference["hubs_by_code"].get(hub_code)
        if hub is None:
            add_error("hub_code", "HUB_NOT_FOUND",
                      f"مركز الانطلاق «{hub_code}» غير مسجّل في البيانات الرئيسية")

        # --------------------------------------------------- الجهات ------
        pickup_code = str(value_of("pickup_facility_code") or "").strip().upper()
        dropoff_code = str(value_of("dropoff_facility_code") or "").strip().upper()
        pickup = reference["facilities_by_code"].get(pickup_code)
        dropoff = reference["facilities_by_code"].get(dropoff_code)
        if pickup is None:
            add_error("pickup_facility_code", "FACILITY_NOT_FOUND",
                      f"جهة الالتقاط «{pickup_code}» غير مسجّلة")
        if dropoff is None:
            add_error("dropoff_facility_code", "FACILITY_NOT_FOUND",
                      f"جهة التسليم «{dropoff_code}» غير مسجّلة")
        if pickup and dropoff and pickup["id"] == dropoff["id"]:
            add_error("dropoff_facility_code", "SAME_FACILITY",
                      "جهة الالتقاط وجهة التسليم متطابقتان")

        if not row.is_valid:
            parsed.append(row)
            continue

        # ------------------------------------------------ الإحداثيات -----
        for label, facility, column in (
            ("الالتقاط", pickup, "pickup_facility_code"),
            ("التسليم", dropoff, "dropoff_facility_code"),
        ):
            if facility["lat"] is None or facility["lon"] is None:
                add_error(column, "MISSING_COORDINATES",
                          f"جهة {label} «{facility['name_ar']}» بلا إحداثيات مسجلة")
            else:
                try:
                    from masar_core.timeutil import validate_coordinates

                    validate_coordinates(
                        facility["lat"], facility["lon"],
                        label=f"إحداثيات جهة {label}",
                    )
                except ValidationError as exc:
                    add_error(column, "BAD_COORDINATES", exc.message)

        # -------------------------------------------- الإعدادات الفعالة --
        if hub["id"] not in settings_cache:
            settings_cache[hub["id"]] = settings_service.effective_for_hub(conn, hub["id"])
        effective = settings_cache[hub["id"]]

        # ------------------------------------------------ نافذة الالتقاط --
        try:
            pickup_clock = parse_time(value_of("pickup_time"), field="موعد الالتقاط")
        except ValidationError as exc:
            add_error("pickup_time", "BAD_TIME", exc.message)
            parsed.append(row)
            continue

        appointment = combine_local(service_date, pickup_clock, timezone_name)
        before = int(effective["pickup_window_before_minutes"])
        after = int(effective["pickup_window_after_minutes"])
        window_from = appointment - dt.timedelta(minutes=before)
        window_to = appointment + dt.timedelta(minutes=after)

        explicit_from = value_of("pickup_window_from")
        explicit_to = value_of("pickup_window_to")
        if explicit_from:
            try:
                window_from = combine_local(
                    service_date, parse_time(explicit_from, field="بداية النافذة"),
                    timezone_name)
            except ValidationError as exc:
                add_error("pickup_window_from", "BAD_TIME", exc.message)
        if explicit_to:
            try:
                window_to = combine_local(
                    service_date, parse_time(explicit_to, field="نهاية النافذة"),
                    timezone_name)
            except ValidationError as exc:
                add_error("pickup_window_to", "BAD_TIME", exc.message)

        if window_to < window_from:
            add_error("pickup_window_to", "INVALID_WINDOW",
                      "نهاية نافذة الالتقاط قبل بدايتها")

        # -------------------------------------------------------- SLA ----
        sla_raw = value_of("sla_time")
        sla_deadline: dt.datetime | None = None
        try:
            text = normalize_digits(str(sla_raw))
            if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", text.strip()):
                sla_clock = parse_time(text, field="الموعد النهائي")
                sla_deadline = combine_local(service_date, sla_clock, timezone_name)
                # ترحيل لليوم التالي فقط في الحالة المنطقية الوحيدة:
                # التقاط مسائي وموعد تسليم في الساعات الأولى بعد منتصف الليل.
                if (
                    sla_deadline <= window_from
                    and pickup_clock.hour >= 12
                    and sla_clock.hour < 12
                ):
                    sla_deadline += dt.timedelta(days=1)
            else:
                from masar_core.timeutil import parse_datetime

                sla_deadline = parse_datetime(
                    sla_raw, field="الموعد النهائي", timezone_name=timezone_name)
        except ValidationError as exc:
            add_error("sla_time", "BAD_SLA", exc.message)

        if sla_deadline is not None and sla_deadline <= window_from:
            add_error("sla_time", "SLA_BEFORE_PICKUP",
                      "الموعد النهائي للتسليم قبل بداية نافذة الالتقاط أو مساوٍ لها")

        # ------------------------------------------------ حقول اختيارية --
        piece_count = 1
        raw_pieces = value_of("piece_count")
        if raw_pieces is not None:
            try:
                piece_count = int(float(normalize_digits(str(raw_pieces))))
                if piece_count < 1:
                    raise ValueError
            except ValueError:
                add_error("piece_count", "BAD_NUMBER", "عدد القطع يجب أن يكون عددًا موجبًا")

        service_type = str(value_of("service_type") or "ROUTINE").strip().upper()
        if service_type not in set(ServiceType):
            add_warning("service_type", "UNKNOWN_SERVICE_TYPE",
                        f"نوع خدمة غير معروف «{service_type}» — استُخدم ROUTINE")
            service_type = "ROUTINE"

        temperature_mode = str(value_of("temperature_mode") or "AMBIENT").strip().upper()
        if temperature_mode not in ("AMBIENT", "CHILLED", "FROZEN", "DEEP_FROZEN", "CONTROLLED"):
            add_warning("temperature_mode", "UNKNOWN_TEMPERATURE",
                        f"نطاق حرارة غير معروف «{temperature_mode}» — استُخدم AMBIENT")
            temperature_mode = "AMBIENT"

        sample_types_raw = value_of("sample_types")
        sample_types = (
            [item.strip() for item in re.split(r"[,،;]", str(sample_types_raw)) if item.strip()]
            if sample_types_raw else []
        )

        if not row.is_valid:
            parsed.append(row)
            continue

        # -------------------------------------- فحص الجدوى المبدئي --------
        feasibility = _prescreen(
            pickup=pickup, dropoff=dropoff, hub=hub,
            window_from=window_from, window_to=window_to,
            sla_deadline=sla_deadline, effective=effective,
        )
        if feasibility is not None:
            add_error(*feasibility)

        row.normalized = {
            "external_reference": value_of("external_reference"),
            "service_date": service_date.isoformat(),
            "service_type": service_type,
            "hub_id": hub["id"],
            "hub_code": hub["code"],
            "region_id": pickup["region_id"],
            "city_id": pickup["city_id"],
            "pickup_facility_id": pickup["id"],
            "pickup_facility_type": pickup["facility_type"],
            "pickup_name": pickup["name_ar"],
            "pickup_lat": pickup["lat"],
            "pickup_lon": pickup["lon"],
            "pickup_address": pickup["address"],
            "pickup_service_minutes": pickup["service_minutes"],
            "pickup_contact_name": value_of("pickup_contact_name") or pickup["contact_name"],
            "pickup_contact_phone": value_of("pickup_contact_phone") or pickup["contact_phone"],
            "pickup_window_from": window_from.isoformat(),
            "pickup_window_to": window_to.isoformat(),
            "dropoff_facility_id": dropoff["id"],
            "dropoff_facility_type": dropoff["facility_type"],
            "dropoff_name": dropoff["name_ar"],
            "dropoff_lat": dropoff["lat"],
            "dropoff_lon": dropoff["lon"],
            "dropoff_address": dropoff["address"],
            "dropoff_service_minutes": dropoff["service_minutes"],
            "dropoff_contact_name": value_of("dropoff_contact_name") or dropoff["contact_name"],
            "dropoff_contact_phone": value_of("dropoff_contact_phone") or dropoff["contact_phone"],
            "sla_deadline": sla_deadline.isoformat() if sla_deadline else None,
            "piece_count": piece_count,
            "sample_types": sample_types,
            "temperature_mode": temperature_mode,
            "notes": value_of("notes"),
        }

        # --------------------------------------------- كشف التكرار --------
        signature = "|".join([
            str(value_of("external_reference") or ""),
            pickup["id"], dropoff["id"], service_date.isoformat(),
            window_from.isoformat(),
        ])
        row.dedupe_key = hashlib.sha256(signature.encode("utf-8")).hexdigest()
        if row.dedupe_key in seen_keys:
            add_error(None, "DUPLICATE_ROW",
                      f"صف مكرر — يطابق الصف رقم {seen_keys[row.dedupe_key]} "
                      "(نفس الجهتين والتاريخ والنافذة)")
        else:
            seen_keys[row.dedupe_key] = row_number

        parsed.append(row)

    return parsed


def _prescreen(
    *, pickup: Any, dropoff: Any, hub: Any,
    window_from: dt.datetime, window_to: dt.datetime,
    sla_deadline: dt.datetime | None, effective: dict[str, Any],
) -> tuple[str | None, str, str] | None:
    """فحص جدوى مبدئي بنفس منطق المحرك — يمنع مرور شحنة مستحيلة إلى التخطيط."""
    if sla_deadline is None:
        return None

    from masar_opt.engine import ShipmentInput, HubInput, VehicleInput, build_problem
    from masar_opt.evaluate import shipment_prescreen

    try:
        problem = build_problem(
            service_date=window_from.date(),
            hubs=[HubInput(
                hub_id=hub["id"], code=hub["code"], name_ar=hub["name_ar"],
                lat=hub["lat"], lon=hub["lon"],
                opens_at=window_from - dt.timedelta(hours=6),
                closes_at=sla_deadline + dt.timedelta(hours=6),
            )],
            shipments=[ShipmentInput(
                shipment_id="probe", reference="probe", hub_id=hub["id"],
                pickup_facility_id=pickup["id"],
                pickup_facility_type=pickup["facility_type"],
                pickup_name=pickup["name_ar"],
                pickup_lat=pickup["lat"], pickup_lon=pickup["lon"],
                pickup_window_from=window_from, pickup_window_to=window_to,
                pickup_service_minutes=pickup["service_minutes"],
                dropoff_facility_id=dropoff["id"],
                dropoff_facility_type=dropoff["facility_type"],
                dropoff_name=dropoff["name_ar"],
                dropoff_lat=dropoff["lat"], dropoff_lon=dropoff["lon"],
                dropoff_service_minutes=dropoff["service_minutes"],
                sla_deadline=sla_deadline,
            )],
            vehicles=[VehicleInput(
                hub_id=hub["id"], label="فحص",
                earliest_start=window_from - dt.timedelta(
                    hours=float(effective["max_shift_hours"])),
                latest_end=sla_deadline + dt.timedelta(hours=1),
                max_shift_minutes=float(effective["max_shift_hours"]) * 60.0,
            )],
            effective_settings=effective,
            fallback_to_estimate=True,
        )
    except Exception:
        return None

    violation = shipment_prescreen(problem.shipments[0], problem, problem.vehicles[0])
    if violation is None:
        return None
    column = "sla_time" if "SLA" in str(violation.reason) else "pickup_time"
    # استبدال الاسم الداخلي للفحص بأسماء الجهات الحقيقية حتى تكون الرسالة
    # مفهومة لمن يصحح الملف
    message = violation.message_ar.replace(
        "الشحنة probe:",
        f"النقل من «{pickup['name_ar']}» إلى «{dropoff['name_ar']}»:",
    )
    return (column, f"INFEASIBLE_{violation.rule}", message)


def _plain(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if value is None:
        return None
    return str(value)


# --------------------------------------------------- عمليات الاستيراد ------

def create_import(
    context: SecurityContext,
    *,
    filename: str,
    content: bytes,
    content_type: str | None,
    storage_key: str,
    is_test_data: bool = False,
    ip_address: str | None = None,
) -> dict[str, Any]:
    """ينشئ سجل استيراد ويقرأ الترويسات ويقترح مطابقة الأعمدة."""
    headers, rows = read_table(content, filename)
    if not rows:
        raise ValidationError("الملف لا يحتوي أي صف بيانات")

    mapping = suggest_mapping(headers)
    reference = f"IMP-{dt.datetime.now(dt.timezone.utc):%Y%m%d}-{uuid.uuid4().hex[:6].upper()}"
    digest = hashlib.sha256(content).hexdigest()

    with transaction(context) as conn:
        duplicate = conn.fetch_one(
            "SELECT reference FROM schedule_imports WHERE sha256 = $1 "
            "AND status <> 'REJECTED' ORDER BY created_at DESC LIMIT 1",
            [digest],
        )
        import_id = conn.fetch_value(
            "INSERT INTO schedule_imports (reference, original_filename, storage_key, "
            "content_type, byte_size, sha256, status, column_mapping, total_rows, "
            "uploaded_by, is_test_data) "
            "VALUES ($1,$2,$3,$4,$5,$6,'MAPPING',$7::jsonb,$8,$9::uuid,$10) "
            "RETURNING id::text",
            [reference, filename, storage_key, content_type, len(content), digest,
             pgwire.Jsonb(mapping), len(rows), context.user_id, is_test_data],
        )
        for offset, raw in enumerate(rows):
            conn.execute(
                "INSERT INTO import_rows (import_id, row_number, raw) "
                "VALUES ($1::uuid, $2, $3::jsonb)",
                [import_id, offset + 2, pgwire.Jsonb({k: _plain(v) for k, v in raw.items()})],
            )
        audit.record(
            conn, context, AuditAction.SCHEDULE_UPLOAD,
            entity_type="schedule_import", entity_id=import_id, entity_label=reference,
            new_value={"filename": filename, "rows": len(rows), "sha256": digest},
            ip_address=ip_address, is_test_data=is_test_data,
        )

    return {
        "id": import_id,
        "reference": reference,
        "headers": headers,
        "mapping": mapping,
        "total_rows": len(rows),
        "preview": [{k: _plain(v) for k, v in row.items()} for row in rows[:20]],
        "duplicate_of": duplicate["reference"] if duplicate else None,
        "template_columns": [
            {
                "key": column.key, "label_ar": column.label_ar,
                "required": column.required, "note_ar": column.note_ar,
            }
            for column in TEMPLATE_COLUMNS
        ],
    }


def validate_import(
    context: SecurityContext,
    import_id: str,
    mapping: dict[str, str | None] | None = None,
) -> dict[str, Any]:
    """يشغّل التحقق الكامل ويحفظ النتائج صفًا صفًا."""
    with transaction(context) as conn:
        record = conn.fetch_one(
            "SELECT id::text AS id, reference, status, column_mapping, is_test_data "
            "FROM schedule_imports WHERE id = $1::uuid",
            [import_id],
        )
        if record is None:
            raise NotFound("سجل الاستيراد غير موجود")
        if record["status"] == ImportStatus.COMMITTED:
            raise Conflict("هذا الاستيراد معتمد بالفعل ولا يمكن إعادة التحقق منه")

        active_mapping = mapping or record["column_mapping"]
        raw_rows = conn.fetch_all(
            "SELECT row_number, raw FROM import_rows WHERE import_id = $1::uuid "
            "ORDER BY row_number",
            [import_id],
        )
        parsed = validate_rows(
            conn, [row["raw"] for row in raw_rows], active_mapping
        )

        valid = sum(1 for row in parsed if row.is_valid)
        duplicates = sum(
            1 for row in parsed
            if any(issue.code == "DUPLICATE_ROW" for issue in row.issues)
        )
        invalid = len(parsed) - valid

        for row in parsed:
            conn.execute(
                "UPDATE import_rows SET normalized = $1::jsonb, is_valid = $2, "
                "errors = $3::jsonb, warnings = $4::jsonb, dedupe_key = $5 "
                "WHERE import_id = $6::uuid AND row_number = $7",
                [
                    pgwire.Jsonb(row.normalized) if row.normalized else None,
                    row.is_valid,
                    pgwire.Jsonb([issue.to_dict() for issue in row.errors]),
                    pgwire.Jsonb([issue.to_dict() for issue in row.warnings]),
                    row.dedupe_key, import_id, row.row_number,
                ],
            )

        dates = sorted({
            row.normalized["service_date"] for row in parsed if row.normalized
        })
        status = (
            ImportStatus.VALIDATED if invalid == 0
            else ImportStatus.PARTIALLY_VALID if valid > 0
            else ImportStatus.REJECTED
        )
        summary = {
            "by_code": _issue_histogram(parsed),
            "dates": dates,
        }
        conn.execute(
            "UPDATE schedule_imports SET status = $1, valid_rows = $2, invalid_rows = $3, "
            "duplicate_rows = $4, column_mapping = $5::jsonb, summary = $6::jsonb, "
            "period_start = $7::date, period_end = $8::date WHERE id = $9::uuid",
            [status, valid, invalid, duplicates, pgwire.Jsonb(active_mapping),
             pgwire.Jsonb(summary), dates[0] if dates else None,
             dates[-1] if dates else None, import_id],
        )

    return {
        "import_id": import_id,
        "status": status,
        "total_rows": len(parsed),
        "valid_rows": valid,
        "invalid_rows": invalid,
        "duplicate_rows": duplicates,
        "dates": dates,
        "issue_summary": summary["by_code"],
        "rows": [
            {
                "row_number": row.row_number,
                "is_valid": row.is_valid,
                "errors": [issue.to_dict() for issue in row.errors],
                "warnings": [issue.to_dict() for issue in row.warnings],
                "normalized": row.normalized or None,
            }
            for row in parsed
        ],
    }


def _issue_histogram(rows: list[ParsedRow]) -> list[dict[str, Any]]:
    counts: dict[tuple[str, str], int] = {}
    labels: dict[tuple[str, str], str] = {}
    for row in rows:
        for issue in row.issues:
            key = (issue.code, issue.severity)
            counts[key] = counts.get(key, 0) + 1
            labels[key] = issue.message_ar
    return [
        {"code": code, "severity": severity, "count": count,
         "sample_message_ar": labels[(code, severity)]}
        for (code, severity), count in sorted(counts.items(), key=lambda x: -x[1])
    ]


def build_error_report_csv(context: SecurityContext, import_id: str) -> bytes:
    """ملف أخطاء قابل للتنزيل (§9 خطوة ١٣)."""
    with session(context) as conn:
        rows = conn.fetch_all(
            "SELECT row_number, raw, errors, warnings FROM import_rows "
            "WHERE import_id = $1::uuid AND (jsonb_array_length(errors) > 0 "
            "OR jsonb_array_length(warnings) > 0) ORDER BY row_number",
            [import_id],
        )
    identity_columns = [
        column.label_ar for column in TEMPLATE_COLUMNS
        if column.key in ("external_reference", "service_date",
                          "pickup_facility_code", "dropoff_facility_code")
    ]
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        ["رقم الصف", *identity_columns, "الخطورة", "العمود", "رمز المشكلة",
         "الوصف", "القيمة الأصلية"]
    )
    for row in rows:
        identity = [str(row["raw"].get(label, "") or "") for label in identity_columns]
        for issue in list(row["errors"]) + list(row["warnings"]):
            column = issue.get("column")
            label = issue.get("column_label_ar") or column or "-"
            original = str(row["raw"].get(label, "") or "")
            writer.writerow([
                row["row_number"], *identity, issue.get("severity", "ERROR"),
                label, issue.get("code", ""), issue.get("message_ar", ""), original,
            ])
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def exclude_rows(context: SecurityContext, import_id: str, row_numbers: list[int]) -> int:
    with transaction(context) as conn:
        result = conn.execute(
            "UPDATE import_rows SET is_excluded = true "
            "WHERE import_id = $1::uuid AND row_number = ANY($2::int[])",
            [import_id, row_numbers],
        )
    return result.rowcount


def commit_import(
    context: SecurityContext,
    import_id: str,
    *,
    skip_invalid: bool = True,
    ip_address: str | None = None,
) -> dict[str, Any]:
    """يحوّل الصفوف الصالحة إلى شحنات بحالة VALIDATED."""
    created = 0
    skipped = 0
    duplicates = 0

    with transaction(context) as conn:
        record = conn.fetch_one(
            "SELECT id::text AS id, reference, status, is_test_data "
            "FROM schedule_imports WHERE id = $1::uuid",
            [import_id],
        )
        if record is None:
            raise NotFound("سجل الاستيراد غير موجود")
        if record["status"] == ImportStatus.COMMITTED:
            raise Conflict("هذا الاستيراد معتمد بالفعل")
        if record["status"] not in (ImportStatus.VALIDATED, ImportStatus.PARTIALLY_VALID):
            raise Conflict(
                f"لا يمكن الاعتماد والحالة {record['status']} — شغّل التحقق أولًا")
        if record["status"] == ImportStatus.PARTIALLY_VALID and not skip_invalid:
            raise Conflict(
                "الملف يحتوي صفوفًا غير صالحة — صحّحها أو فعّل استبعاد الصفوف غير الصالحة")

        rows = conn.fetch_all(
            "SELECT id::text AS id, row_number, normalized, dedupe_key FROM import_rows "
            "WHERE import_id = $1::uuid AND is_valid AND NOT is_excluded "
            "ORDER BY row_number",
            [import_id],
        )

        sequence = 0
        for row in rows:
            data = row["normalized"]
            sequence += 1
            reference = f"{record['reference'][4:]}-{row['row_number']:05d}"
            # نقطة حفظ لكل صف: خرق تفرّد في صف واحد يُرجع ذلك الصف وحده،
            # ولا يُجهض المعاملة الأم ولا بقية الاستيراد.
            try:
                with conn.transaction():
                    shipment_id = conn.fetch_value(
                        """
                        INSERT INTO shipments (
                            reference, external_reference, request_kind, service_type, status,
                            region_id, city_id, hub_id,
                            pickup_facility_id, pickup_facility_type, pickup_contact_name,
                            pickup_contact_phone, pickup_address, pickup_lat, pickup_lon,
                            pickup_window_from, pickup_window_to, pickup_service_minutes,
                            dropoff_facility_id, dropoff_facility_type, dropoff_contact_name,
                            dropoff_contact_phone, dropoff_address, dropoff_lat, dropoff_lon,
                            sla_deadline, dropoff_service_minutes,
                            piece_count, sample_types, temperature_mode, service_date,
                            import_id, import_row_number, notes, dedupe_key, is_test_data
                        ) VALUES (
                            $1,$2,'SCHEDULED',$3,'VALIDATED',
                            $4::uuid,$5::uuid,$6::uuid,
                            $7::uuid,$8,$9,$10,$11,$12,$13,
                            $14::timestamptz,$15::timestamptz,$16,
                            $17::uuid,$18,$19,$20,$21,$22,$23,
                            $24::timestamptz,$25,
                            $26,$27::text[],$28,$29::date,
                            $30::uuid,$31,$32,$33,$34
                        ) RETURNING id::text
                        """,
                        [
                            reference, data.get("external_reference"), data["service_type"],
                            data["region_id"], data["city_id"], data["hub_id"],
                            data["pickup_facility_id"], data["pickup_facility_type"],
                            data.get("pickup_contact_name"), data.get("pickup_contact_phone"),
                            data.get("pickup_address"), data["pickup_lat"], data["pickup_lon"],
                            data["pickup_window_from"], data["pickup_window_to"],
                            data["pickup_service_minutes"],
                            data["dropoff_facility_id"], data["dropoff_facility_type"],
                            data.get("dropoff_contact_name"), data.get("dropoff_contact_phone"),
                            data.get("dropoff_address"), data["dropoff_lat"], data["dropoff_lon"],
                            data["sla_deadline"], data["dropoff_service_minutes"],
                            data["piece_count"], data.get("sample_types") or [],
                            data["temperature_mode"], data["service_date"],
                            import_id, row["row_number"], data.get("notes"),
                            row["dedupe_key"], record["is_test_data"],
                        ],
                    )
                    conn.execute(
                        "UPDATE import_rows SET shipment_id = $1::uuid WHERE id = $2::uuid",
                        [shipment_id, row["id"]],
                    )
                created += 1
            except pgwire.UniqueViolation:
                duplicates += 1
                skipped += 1

        conn.execute(
            "UPDATE schedule_imports SET status = 'COMMITTED', committed_by = $1::uuid, "
            "committed_at = now() WHERE id = $2::uuid",
            [context.user_id, import_id],
        )
        audit.record(
            conn, context, AuditAction.SCHEDULE_COMMIT,
            entity_type="schedule_import", entity_id=import_id,
            entity_label=record["reference"],
            new_value={"created": created, "skipped": skipped, "duplicates": duplicates},
            ip_address=ip_address, is_test_data=record["is_test_data"],
        )

    return {
        "import_id": import_id,
        "created_shipments": created,
        "skipped_rows": skipped,
        "duplicate_shipments": duplicates,
    }


def get_import(context: SecurityContext, import_id: str) -> dict[str, Any]:
    with session(context) as conn:
        record = conn.fetch_one(
            "SELECT i.id::text AS id, i.reference, i.original_filename, i.status, "
            "i.total_rows, i.valid_rows, i.invalid_rows, i.duplicate_rows, "
            "i.column_mapping, i.summary, i.period_start, i.period_end, i.created_at, "
            "i.committed_at, u.full_name AS uploaded_by_name "
            "FROM schedule_imports i LEFT JOIN users u ON u.id = i.uploaded_by "
            "WHERE i.id = $1::uuid",
            [import_id],
        )
        if record is None:
            raise NotFound("سجل الاستيراد غير موجود")
        rows = conn.fetch_all(
            "SELECT row_number, is_valid, is_excluded, errors, warnings, normalized, raw "
            "FROM import_rows WHERE import_id = $1::uuid ORDER BY row_number",
            [import_id],
        )
    return {"import": dict(record), "rows": [dict(row) for row in rows]}


def list_imports(context: SecurityContext, limit: int = 50) -> list[Any]:
    with session(context) as conn:
        return conn.fetch_all(
            "SELECT i.id::text AS id, i.reference, i.original_filename, i.status, "
            "i.total_rows, i.valid_rows, i.invalid_rows, i.period_start, i.period_end, "
            "i.created_at, i.committed_at, u.full_name AS uploaded_by_name "
            "FROM schedule_imports i LEFT JOIN users u ON u.id = i.uploaded_by "
            f"ORDER BY i.created_at DESC LIMIT {int(limit)}"
        )
